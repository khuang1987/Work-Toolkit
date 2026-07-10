import openpyxl
from openpyxl import load_workbook
# 新增导入，用于清除单元格格式
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment  
import os
from datetime import datetime
import sys  # 新增sys导入

def process_excel_file(file_path, callback=None):
    # 如果没有提供回调函数，使用默认的print
    if callback is None:
        callback = print
    
    # 打开Excel文件并输出文件名
    callback(f"打开文件: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active

    # 检查K列或L列第一行是否已有数据（非空），若是，则返回False
    if ws.cell(row=1, column=11).value not in (None, '') or ws.cell(row=1, column=12).value not in (None, ''):
        callback("检测到已处理文件，跳过处理")
        return False
    
    # 清除整个工作表的底色、框线，并将所有数据字体改为宋体 11号，单元格内容上下左右居中且高度自动
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border(
                left=Side(border_style=None),
                right=Side(border_style=None),
                top=Side(border_style=None),
                bottom=Side(border_style=None)
            )
            cell.font = Font(name="SimSun", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)  # 新增设置居中和自动高度
    
    # 设置K列和L列的标题
    ws.cell(row=1, column=11, value='合格数')
    ws.cell(row=1, column=12, value='不合格数')
    
    # 遍历H列并处理数据
    modified_rows = 0
    for row in range(2, ws.max_row + 1):
        h_cell = ws.cell(row=row, column=8)
        result = str(h_cell.value).strip() if h_cell.value else ''
        
        k_value = 0
        l_value = 0
        
        if '合格数：' in result:
            parts = result.split('/')
            # 处理合格数
            hege_part = parts[0].strip()
            if '合格数：' in hege_part:
                hege_str = hege_part.split('：')[-1].strip()
                k_value = int(hege_str) if hege_str.isdigit() else 0
            
            # 处理不合格数
            if len(parts) > 1:
                buhege_part = parts[1].strip()
                if '不合格数：' in buhege_part:
                    buhege_str = buhege_part.split('：')[-1].strip()
                    l_value = int(buhege_str) if buhege_str.isdigit() else 0
        else:
            if result == '合格':
                k_value = 1
            elif result == '不合格':
                l_value = 1
        
        # 写入结果
        ws.cell(row=row, column=11, value=k_value)  # K列
        ws.cell(row=row, column=12, value=l_value)  # L列
        modified_rows += 1
    
    # 输出修改的行数
    callback(f"共修改了 {modified_rows} 行数据")
    
    # 保存文件
    wb.save(file_path)
    callback(f"保存文件: {file_path}")
    return True

def main(callback=None):
    # 如果没有提供回调函数，使用默认的print
    if callback is None:
        callback = print
        
    username = os.getlogin()  # 获取当前登录用户名
    # 获取文件夹中的所有Excel文件
    folder_path = rf'C:\Users\{username}\OneDrive - Medtronic PLC\General - CZ Production\POWER BI 数据源 V2\70-SFC导出数据\班组合格率数据'
    files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    files.sort(key=lambda x: os.path.getctime(os.path.join(folder_path, x)), reverse=True)

    # 遍历所有文件并处理
    for file in files:
        file_path = os.path.join(folder_path, file)
        if not process_excel_file(file_path, callback):
            callback("检测到已处理文件，结束处理")
            break

    callback("所有文件处理完成")

class FileAlreadyProcessedError(Exception):
    """当文件已经被处理过时抛出此异常"""
    pass

if __name__ == "__main__":
    main()