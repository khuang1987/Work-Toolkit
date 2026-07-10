import gc
import json
import logging
import os
import sys
import threading
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from tkinter import filedialog, messagebox
import tkinter as tk

import customtkinter as ctk
import pandas as pd
import pythoncom
import win32com.client as win32


if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent.parent

CONFIG_PATH = BASE_DIR / "config.json"
LOG_PATH = BASE_DIR / "transfer.log"
UI_FONT = "Microsoft YaHei"

DEFAULT_CONFIG = {
    "source_file": "",
    "source_dir": "",
    "source_keyword": "巡线记录",
    "target_file": "",
    "sheet_name": "加工中心",
    "target_sheet_name": "主管",
    "backup_csv_dir": "publish",
    "table_name": "表1",
    "preferred_registrant": "纪磊",
    "preferred_registrant_limit": 5,
    "min_record_limit": 10,
    "max_record_limit": 13,
    "default_registrant": "黄凯",
    "default_area": "CZ 加工中心",
    "default_process": "数控铣",
    "keep_excel_open": True
}

SOURCE_REQUIRED_COLUMNS = [
    "日期",
    "区域",
    "设备编号",
    "登记人",
]

COLUMN_ALIASES = {
    "产品图号": ["产品图号", "产品号"],
    "问题类型": ["问题类型", "问题类型\n(必填）", "问题类型\n(必填)", "问题类型(必填）", "问题类型(必填)"],
    "检查正常?": ["检查正常?", "检查正常"],
    "工单号/批次号": ["工单号/批次号"],
    "责任人": ["责任人"],
    "首件是否在中巡": ["首件是否在中巡", "首件是否在中差"],
    "达到警报线的下一件是否有调回中差": ["达到警报线的下一件是否有调回中差"],
    "员工正确打开岗位所有文件": ["员工正确打开岗位所有文件"],
    "清场确认": ["清场确认"],
    "事件描述": ["事件描述"],
    "班长": ["班长"],
}


def configure_logging():
    root_logger = logging.getLogger()
    if root_logger.handlers:
        return

    root_logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)

    file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
    file_handler.setFormatter(formatter)

    root_logger.addHandler(stream_handler)
    root_logger.addHandler(file_handler)


configure_logging()


def load_config():
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as file:
                loaded = json.load(file)
        except Exception as exc:
            logging.error("加载配置失败: %s", exc)
            loaded = {}
    else:
        loaded = {}

    config = DEFAULT_CONFIG.copy()
    config.update(loaded)
    return config


def save_config(config):
    stored = DEFAULT_CONFIG.copy()
    stored.update(config)
    with open(CONFIG_PATH, "w", encoding="utf-8") as file:
        json.dump(stored, file, ensure_ascii=False, indent=2)


CONFIG = load_config()


def resolve_publish_dir(config):
    publish_dir = config.get("backup_csv_dir", DEFAULT_CONFIG["backup_csv_dir"])
    publish_path = Path(publish_dir)
    if not publish_path.is_absolute():
        publish_path = BASE_DIR / publish_path
    publish_path.mkdir(parents=True, exist_ok=True)
    return publish_path


def require_path(path_value, label):
    if not path_value:
        raise ValueError(f"请先选择{label}")

    path = Path(path_value)
    if not path.exists():
        raise FileNotFoundError(f"{label}不存在: {path}")
    return path


def find_source_file(config):
    source_file_value = (config.get("source_file") or "").strip()
    if source_file_value:
        source_path = Path(source_file_value)
        if source_path.is_file():
            return source_path
        if source_path.is_dir():
            config = config.copy()
            config["source_dir"] = str(source_path)

    source_dir_value = (config.get("source_dir") or "").strip()
    if not source_dir_value:
        raise ValueError("请先选择采集文件，或配置采集目录")

    source_dir = require_path(source_dir_value, "采集目录")
    if not source_dir.is_dir():
        raise NotADirectoryError(f"采集目录不是文件夹: {source_dir}")

    keyword = (config.get("source_keyword") or DEFAULT_CONFIG["source_keyword"]).strip()
    candidates = [
        path for path in source_dir.glob("*.xlsx")
        if keyword in path.name and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(f"采集目录中未找到文件名包含“{keyword}”的 xlsx 文件: {source_dir}")

    selected = max(candidates, key=lambda path: path.stat().st_mtime)
    logging.info("自动匹配采集文件: %s", selected)
    return selected


def get_target_week_range(today=None):
    today = today or datetime.now()
    weekday = today.weekday()

    if weekday >= 5:
        days_since_friday = weekday - 4
        week_end = today - timedelta(days=days_since_friday)
        week_start = week_end - timedelta(days=4)
    else:
        current_week_start = today - timedelta(days=weekday)
        week_start = current_week_start - timedelta(days=7)
        week_end = week_start + timedelta(days=4)

    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_end.replace(hour=23, minute=59, second=59, microsecond=999999)
    return week_start, week_end


def normalize_source_dataframe(df):
    if "日期" not in df.columns:
        raise KeyError("源文件缺少“日期”列")

    df = df.copy()
    df["日期"] = pd.to_datetime(df["日期"], errors="coerce")

    missing_columns = [column for column in SOURCE_REQUIRED_COLUMNS if column not in df.columns]
    if missing_columns:
        raise KeyError(f"源文件缺少必要列: {', '.join(missing_columns)}")

    return df


def get_first_available_value(record, aliases, default=""):
    for alias in aliases:
        if alias in record.index:
            return record.get(alias, default)
    return default


def workbook_matches_target(workbook, target_path):
    workbook_name = Path(str(workbook.Name)).name.lower()
    target_name = Path(str(target_path)).name.lower()
    if workbook_name == target_name:
        return True

    full_name = str(workbook.FullName).replace("/", "\\").lower()
    return full_name.endswith("\\" + target_name)


def find_open_target_workbook(target_path, include_objects=False):
    initialized = False
    try:
        if not include_objects:
            pythoncom.CoInitialize()
            initialized = True
        excel = win32.GetActiveObject("Excel.Application")
    except Exception:
        return None

    try:
        for workbook in excel.Workbooks:
            if workbook_matches_target(workbook, target_path):
                result = {
                    "name": str(workbook.Name),
                    "full_name": str(workbook.FullName),
                }
                if include_objects:
                    result["excel"] = excel
                    result["workbook"] = workbook
                return result
    except Exception as exc:
        logging.warning("检查目标文件打开状态失败: %s", exc)
    finally:
        if initialized and not include_objects:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

    return None


def extract_last_week_records(config):
    source_file = find_source_file(config)
    sheet_name = config.get("sheet_name", DEFAULT_CONFIG["sheet_name"])

    logging.info("读取源文件: %s", source_file)
    logging.info("读取工作表: %s", sheet_name)

    df = pd.read_excel(source_file, sheet_name=sheet_name)
    df = normalize_source_dataframe(df)

    week_start, week_end = get_target_week_range()
    logging.info(
        "筛选上周工作日日期范围: %s 到 %s",
        week_start.strftime("%Y-%m-%d"),
        week_end.strftime("%Y-%m-%d"),
    )

    mask = (df["日期"] >= week_start) & (df["日期"] <= week_end)
    last_week_records = df.loc[mask].copy()
    last_week_records = last_week_records[last_week_records[SOURCE_REQUIRED_COLUMNS].notna().all(axis=1)]

    logging.info("筛选后记录数: %s", len(last_week_records))
    if last_week_records.empty:
        return last_week_records

    registrant_counts = last_week_records["登记人"].value_counts()
    logging.info("登记人分布:")
    for name, count in registrant_counts.items():
        logging.info("  %s: %s 条", name, count)

    preferred_name = config.get("preferred_registrant", DEFAULT_CONFIG["preferred_registrant"])
    preferred_limit = int(config.get("preferred_registrant_limit", DEFAULT_CONFIG["preferred_registrant_limit"]))
    min_limit = int(config.get("min_record_limit", DEFAULT_CONFIG["min_record_limit"]))
    max_limit = int(config.get("max_record_limit", DEFAULT_CONFIG["max_record_limit"]))
    if min_limit > max_limit:
        raise ValueError("最少记录数不能大于最多记录数")

    preferred_records = last_week_records[last_week_records["登记人"] == preferred_name]
    other_records = last_week_records[last_week_records["登记人"] != preferred_name]

    selected_preferred = preferred_records.head(preferred_limit)
    if len(last_week_records) <= max_limit:
        target_count = len(last_week_records)
    else:
        target_count = max_limit

    if target_count < min_limit and len(last_week_records) >= min_limit:
        target_count = min_limit

    other_needed = max(0, target_count - len(selected_preferred))
    selected_other = other_records.head(other_needed)

    selected_records = pd.concat([selected_preferred, selected_other], ignore_index=True)
    logging.info(
        "已选记录: %s 条，其中 %s %s 条，其余 %s 条",
        len(selected_records),
        preferred_name,
        len(selected_preferred),
        len(selected_other),
    )

    for idx, record in selected_records.iterrows():
        logging.info(
            "  %s. %s - %s - 设备%s",
            idx + 1,
            record["日期"].strftime("%Y-%m-%d"),
            record["登记人"],
            record["设备编号"],
        )

    return selected_records


def fill_y_if_empty(value):
    if pd.isna(value) or (isinstance(value, str) and not value.strip()):
        return "Y"
    return value


def fill_na_if_empty(value):
    if pd.isna(value) or (isinstance(value, str) and not value.strip()):
        return "N/A"
    return value


def to_excel_date(value):
    if pd.isna(value):
        return ""
    if hasattr(value, "to_pydatetime"):
        value = value.to_pydatetime()
    return value.replace(hour=0, minute=0, second=0, microsecond=0)


def to_com_value(value):
    if value is None:
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().replace(tzinfo=None)
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime().replace(tzinfo=None)
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def format_preview_date(value):
    if pd.isna(value) or not value:
        return ""
    if isinstance(value, str):
        return value
    return f"{value.month}/{value.day}/{value.year}"


def build_output_dataframe(records_df, config):
    output_rows = []
    default_registrant = config.get("default_registrant", DEFAULT_CONFIG["default_registrant"])
    default_area = config.get("default_area", DEFAULT_CONFIG["default_area"])
    default_process = config.get("default_process", DEFAULT_CONFIG["default_process"])

    for _, record in records_df.iterrows():
        work_order_batch = record.get("工单号/批次号", "")
        if work_order_batch == "":
            work_order_batch = get_first_available_value(record, COLUMN_ALIASES["工单号/批次号"], "")
        if pd.notna(work_order_batch):
            work_order_batch = str(work_order_batch).strip()
        else:
            work_order_batch = ""
        if not work_order_batch:
            work_order_batch = "N/A"

        check_status = get_first_available_value(record, COLUMN_ALIASES["检查正常?"], "")
        if pd.notna(check_status) and str(check_status).strip() == "正常":
            p_value = d_value = c_value = a_value = "N/A"
        else:
            p_value = "立即整改"
            d_value = "已完成"
            c_value = "继续检查"
            a_value = "班会宣导"

        output_rows.append(
            {
                "登记人": default_registrant,
                "日期": to_excel_date(record["日期"]),
                "区域": default_area,
                "工序": default_process,
                "设备编号": record.get("设备编号", ""),
                "工单号/批次号": work_order_batch,
                "产品图号": fill_na_if_empty(get_first_available_value(record, COLUMN_ALIASES["产品图号"], "")),
                "责任人": fill_na_if_empty(get_first_available_value(record, COLUMN_ALIASES["责任人"], "")),
                "首件是否在中巡": fill_y_if_empty(get_first_available_value(record, COLUMN_ALIASES["首件是否在中巡"], "")),
                "达到警报线的下一件是否有调回中差": fill_y_if_empty(get_first_available_value(record, COLUMN_ALIASES["达到警报线的下一件是否有调回中差"], "")),
                "5S/DHR": fill_y_if_empty(record.get("5S/DHR", "")),
                "员工正确打开岗位所有文件": fill_y_if_empty(get_first_available_value(record, COLUMN_ALIASES["员工正确打开岗位所有文件"], "")),
                "清场确认": fill_y_if_empty(get_first_available_value(record, COLUMN_ALIASES["清场确认"], "")),
                "事件描述": fill_na_if_empty(get_first_available_value(record, COLUMN_ALIASES["事件描述"], "")),
                "P": p_value,
                "D": d_value,
                "C": c_value,
                "A": a_value,
                "问题类型": fill_na_if_empty(get_first_available_value(record, COLUMN_ALIASES["问题类型"], "")),
                "检查正常?": check_status if pd.notna(check_status) else "",
                "班长": fill_na_if_empty(get_first_available_value(record, COLUMN_ALIASES["班长"], "")),
                "是否录入OQW系统": None,
                "ID": None,
                "提出项次": None,
            }
        )

    return pd.DataFrame(output_rows)


DEDUP_COLUMNS = [
    "日期",
    "区域",
    "工序",
    "设备编号",
    "工单号/批次号",
    "产品图号",
    "事件描述",
    "问题类型",
    "检查正常?",
]

TARGET_COLUMN_ALIASES = {
    "设备编号": ["设备编号", "机床号"],
    "产品图号": ["产品图号", "产品号"],
    "首件是否在中巡": ["首件是否在中巡", "首件是否在中差"],
    "检查正常?": ["检查正常?", "检查正常"],
}


def get_column_value(row, column_name):
    aliases = TARGET_COLUMN_ALIASES.get(column_name, [column_name])
    for alias in aliases:
        if alias in row.index:
            return row.get(alias)
    return None


def normalize_key_value(value):
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        if value.tzinfo is not None:
            value = value.tz_convert("Asia/Shanghai").tz_localize(None)
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone(timedelta(hours=8))).replace(tzinfo=None)
        return value.strftime("%Y-%m-%d")
    if all(hasattr(value, attr) for attr in ("year", "month", "day")):
        return f"{value.year:04d}-{value.month:02d}-{value.day:02d}"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    normalized = str(value).strip()
    if normalized.upper() in {"N/A", "NA", "NONE", "NAN"}:
        return ""
    if normalized.endswith(".0") and normalized[:-2].isdigit():
        return normalized[:-2]
    return normalized


def make_dedup_key(row, date_offset_days=0):
    values = []
    for column in DEDUP_COLUMNS:
        value = get_column_value(row, column)
        if column == "日期" and date_offset_days and isinstance(value, (pd.Timestamp, datetime)):
            value = value + timedelta(days=date_offset_days)
        values.append(normalize_key_value(value))
    return tuple(values)


def deduplicate_output_rows(output_df):
    if output_df.empty:
        return output_df, 0

    deduped = output_df.copy()
    deduped["_dedup_key"] = deduped.apply(make_dedup_key, axis=1)
    before_count = len(deduped)
    deduped = deduped.drop_duplicates("_dedup_key", keep="first").drop(columns=["_dedup_key"])
    removed_count = before_count - len(deduped)
    if removed_count:
        logging.info("本批次去重已移除 %s 条重复记录", removed_count)
    return deduped.reset_index(drop=True), removed_count


def read_existing_target_keys(config):
    target_file = require_path(config.get("target_file", ""), "目标文件")
    target_sheet_name = config.get("target_sheet_name", DEFAULT_CONFIG["target_sheet_name"])
    try:
        target_df = pd.read_excel(target_file, sheet_name=target_sheet_name)
    except Exception as exc:
        logging.warning("直接读取目标文件用于去重失败，尝试读取已打开的 Excel: %s", exc)
        target_df = read_open_target_dataframe(config)
        if target_df is None:
            logging.warning("读取已打开的目标 Excel 失败，将跳过目标去重")
            return set()

    target_df = target_df.dropna(how="all")
    if target_df.empty:
        return set()

    keys = set()
    for _, row in target_df.iterrows():
        keys.add(make_dedup_key(row))
        keys.add(make_dedup_key(row, date_offset_days=1))
    keys.discard(tuple("" for _ in DEDUP_COLUMNS))
    logging.info("已读取目标文件去重基准: %s 条", len(keys))
    return keys


def read_open_target_dataframe(config):
    target_file = require_path(config.get("target_file", ""), "目标文件")
    target_sheet_name = config.get("target_sheet_name", DEFAULT_CONFIG["target_sheet_name"])
    table_name = config.get("table_name", DEFAULT_CONFIG["table_name"])
    open_workbook = find_open_target_workbook(target_file, include_objects=True)
    if not open_workbook:
        return None

    workbook = open_workbook["workbook"]
    try:
        worksheet = workbook.Worksheets[target_sheet_name]
        try:
            values = worksheet.ListObjects(table_name).Range.Value
        except Exception:
            values = worksheet.UsedRange.Value

        if not values:
            return pd.DataFrame()
        if not isinstance(values, tuple):
            values = ((values,),)

        rows = [list(row) if isinstance(row, tuple) else [row] for row in values]
        if len(rows) < 2:
            return pd.DataFrame()

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        data_rows = rows[1:]
        return pd.DataFrame(data_rows, columns=headers, dtype=object)
    except Exception as exc:
        logging.warning("从已打开的目标 Excel 读取去重数据失败: %s", exc)
        return None
    finally:
        try:
            del workbook
        except Exception:
            pass


def filter_existing_output_rows(output_df, config):
    if output_df.empty:
        return output_df, 0

    existing_keys = read_existing_target_keys(config)
    if not existing_keys:
        return output_df, 0

    filtered = output_df.copy()
    filtered["_dedup_key"] = filtered.apply(make_dedup_key, axis=1)
    before_count = len(filtered)
    filtered = filtered[~filtered["_dedup_key"].isin(existing_keys)].drop(columns=["_dedup_key"])
    removed_count = before_count - len(filtered)
    if removed_count:
        logging.info("目标文件去重已跳过 %s 条已存在记录", removed_count)
    return filtered.reset_index(drop=True), removed_count


def save_backup_csv(output_df, config):
    publish_dir = resolve_publish_dir(config)
    csv_path = publish_dir / "待粘贴信息.csv"
    csv_df = output_df.copy()
    if "日期" in csv_df.columns:
        csv_df["日期"] = csv_df["日期"].apply(format_preview_date)
    csv_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    logging.info("CSV备份已保存: %s", csv_path)
    return csv_path


def find_last_valid_row(worksheet, table):
    table_range = table.Range
    start_row = table_range.Row
    end_row = start_row + table_range.Rows.Count - 1

    for check_row in range(end_row, start_row - 1, -1):
        values = [
            worksheet.Cells(check_row, 1).Value,
            worksheet.Cells(check_row, 2).Value,
            worksheet.Cells(check_row, 5).Value,
        ]
        if all(value is not None and str(value).strip() for value in values):
            return check_row

    return start_row


def write_excel_via_com(output_df, config):
    target_file = require_path(config.get("target_file", ""), "目标文件")
    target_sheet_name = config.get("target_sheet_name", DEFAULT_CONFIG["target_sheet_name"])
    table_name = config.get("table_name", DEFAULT_CONFIG["table_name"])
    keep_excel_open = bool(config.get("keep_excel_open", True))

    logging.info("开始写入目标文件: %s", target_file)

    excel = None
    workbook = None
    created_excel = False

    try:
        pythoncom.CoInitialize()

        open_workbook = find_open_target_workbook(target_file, include_objects=True)
        if open_workbook:
            excel = open_workbook["excel"]
            workbook = open_workbook["workbook"]
            created_excel = False
            logging.info("目标文件已打开，将直接写入当前工作簿: %s", open_workbook["full_name"])

        if excel is None:
            import win32com.client

            excel = win32com.client.DispatchEx("Excel.Application")
            created_excel = True
            excel.Visible = True
            excel.DisplayAlerts = False
            logging.info("已创建新的 Excel 实例")

        if workbook is None:
            workbook = excel.Workbooks.Open(str(target_file))
            logging.info("已打开目标文件")

        worksheet = workbook.Worksheets[target_sheet_name]
        table = worksheet.ListObjects(table_name)
        last_valid_row = find_last_valid_row(worksheet, table)
        first_new_row = last_valid_row + 1

        logging.info("将从第 %s 行开始写入", first_new_row)

        write_columns = [
            "登记人",
            "日期",
            "区域",
            "工序",
            "设备编号",
            "工单号/批次号",
            "产品图号",
            "责任人",
            "首件是否在中巡",
            "达到警报线的下一件是否有调回中差",
            "5S/DHR",
            "员工正确打开岗位所有文件",
            "清场确认",
            "事件描述",
            "P",
            "D",
            "C",
            "A",
            "问题类型",
            "检查正常?",
            "班长",
            "是否录入OQW系统",
            "ID",
            "提出项次",
        ]

        data_write_columns = write_columns[:21]
        written_count = 0
        for row_offset, (_, record) in enumerate(output_df.iterrows()):
            target_row = first_new_row + row_offset
            for col_offset, column_name in enumerate(data_write_columns, start=1):
                worksheet.Cells(target_row, col_offset).Value = to_com_value(record.get(column_name))
            worksheet.Cells(target_row, 2).NumberFormat = "m/d/yyyy"
            written_count += 1

        logging.info("已逐格写入第 %s 到 %s 行，共 %s 条", first_new_row, first_new_row + written_count - 1, written_count)

        workbook.Save()
        logging.info("Excel 写入完成，共 %s 条", written_count)

        if keep_excel_open:
            excel.Visible = True
            worksheet.Activate()
            worksheet.Cells(first_new_row, 1).Select()
            excel.ActiveWindow.ScrollRow = max(1, first_new_row - 5)
        elif created_excel:
            workbook.Close(SaveChanges=True)
            excel.Quit()

        return True
    except Exception as exc:
        logging.exception("COM写入失败: %s", exc)
        return False
    finally:
        if workbook is not None:
            del workbook
        if excel is not None:
            del excel
        gc.collect()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def write_excel_via_openpyxl(output_df, config):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    target_file = require_path(config.get("target_file", ""), "目标文件")
    publish_dir = resolve_publish_dir(config)
    output_file = publish_dir / f"巡线记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    logging.info("生成备份 Excel: %s", output_file)

    wb = Workbook()
    ws = wb.active
    ws.title = "主管巡线记录"

    headers = list(output_df.columns[:-3]) + ["是否录入OQW系统", "ID", "提出项次"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (_, record) in enumerate(output_df.iterrows(), start=2):
        for col_idx, column_name in enumerate(headers, start=1):
            value = record.get(column_name)
            if isinstance(value, pd.Timestamp):
                value = value.to_pydatetime().replace(tzinfo=None)
            elif pd.isna(value):
                value = None
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.cell(row=row_idx, column=2).number_format = "m/d/yyyy"

    wb.save(output_file)
    logging.info("已生成备份 Excel: %s", output_file)
    try:
        os.startfile(str(output_file))
    except Exception as exc:
        logging.warning("打开备份 Excel 失败: %s", exc)
    try:
        os.startfile(str(target_file))
    except Exception as exc:
        logging.warning("打开目标 Excel 失败: %s", exc)
    return str(output_file)


def write_records_to_excel(output_df, config):
    if write_excel_via_com(output_df, config):
        return {
            "mode": "target",
            "excel_path": str(require_path(config.get("target_file", ""), "目标文件")),
        }

    logging.warning("目标文件写入失败，改为生成备份 Excel。")
    backup_path = write_excel_via_openpyxl(output_df, config)
    return {
        "mode": "backup",
        "excel_path": backup_path,
    }


def build_runtime_config(config_override=None):
    config = DEFAULT_CONFIG.copy()
    config.update(CONFIG)
    if config_override:
        config.update(config_override)
    return config


def prepare_transfer(config_override=None):
    config = build_runtime_config(config_override)

    source_file = find_source_file(config)
    target_file = require_path(config.get("target_file", ""), "目标文件")
    logging.info("当前采集文件: %s", source_file)
    logging.info("当前目标文件: %s", target_file)

    records = extract_last_week_records(config)
    if records.empty:
        logging.warning("没有找到可导入的数据")
        return {
            "success": False,
            "message": "没有找到上周工作日的可导入记录。",
            "record_count": 0,
        }

    output_df = build_output_dataframe(records, config)
    output_df, source_duplicate_count = deduplicate_output_rows(output_df)
    output_df, target_duplicate_count = filter_existing_output_rows(output_df, config)
    if output_df.empty:
        logging.warning("上周工作日记录去重后没有新数据")
        return {
            "success": False,
            "message": "上周工作日记录去重后没有新数据需要写入。",
            "record_count": 0,
            "source_duplicate_count": source_duplicate_count,
            "target_duplicate_count": target_duplicate_count,
        }

    return {
        "success": True,
        "message": f"预览已生成，共 {len(output_df)} 条新记录。",
        "record_count": len(output_df),
        "records": records,
        "output_df": output_df,
        "config": config,
        "source_file": str(source_file),
        "target_file": str(target_file),
        "source_duplicate_count": source_duplicate_count,
        "target_duplicate_count": target_duplicate_count,
    }


def run_transfer(config_override=None, prepared_output_df=None):
    if prepared_output_df is None:
        prepared = prepare_transfer(config_override)
        if not prepared["success"]:
            return prepared
        config = prepared["config"]
        output_df = prepared["output_df"]
    else:
        config = build_runtime_config(config_override)
        output_df = prepared_output_df

    csv_path = save_backup_csv(output_df, config)
    write_result = write_records_to_excel(output_df, config)
    success = bool(write_result)

    if success and write_result["mode"] == "target":
        message = f"目标文件写入完成，共 {len(output_df)} 条记录。\n文件: {write_result['excel_path']}\nCSV 备份: {csv_path}"
    elif success:
        message = f"目标文件写入失败，已生成备份 Excel，共 {len(output_df)} 条记录。\n文件: {write_result['excel_path']}\nCSV 备份: {csv_path}"
    else:
        message = f"Excel 写入失败，但已生成 CSV 备份。\nCSV 备份: {csv_path}"

    return {
        "success": success,
        "message": message,
        "record_count": len(output_df),
        "csv_path": str(csv_path),
        "excel_path": write_result["excel_path"] if write_result else "",
        "write_mode": write_result["mode"] if write_result else "",
    }


class TransferApp:
    def __init__(self, root):
        self.root = root
        self.root.title("巡线记录导入")
        self.root.geometry("1060x720")
        self.root.minsize(960, 640)
        self.root.configure(fg_color="#f5f5f7")

        self.source_var = tk.StringVar(value=CONFIG.get("source_file", ""))
        self.source_dir_var = tk.StringVar(value=CONFIG.get("source_dir", ""))
        self.source_mode_var = tk.StringVar(value="folder" if CONFIG.get("source_dir", "") else "file")
        self.source_keyword_var = tk.StringVar(value=CONFIG.get("source_keyword", DEFAULT_CONFIG["source_keyword"]))
        self.target_var = tk.StringVar(value=CONFIG.get("target_file", ""))
        self.sheet_var = tk.StringVar(value=CONFIG.get("sheet_name", DEFAULT_CONFIG["sheet_name"]))
        self.target_sheet_var = tk.StringVar(value=CONFIG.get("target_sheet_name", DEFAULT_CONFIG["target_sheet_name"]))
        self.table_name_var = tk.StringVar(value=CONFIG.get("table_name", DEFAULT_CONFIG["table_name"]))
        self.preferred_var = tk.StringVar(value=CONFIG.get("preferred_registrant", DEFAULT_CONFIG["preferred_registrant"]))
        self.preferred_limit_var = tk.StringVar(value=str(CONFIG.get("preferred_registrant_limit", DEFAULT_CONFIG["preferred_registrant_limit"])))
        self.min_limit_var = tk.StringVar(value=str(CONFIG.get("min_record_limit", DEFAULT_CONFIG["min_record_limit"])))
        self.max_limit_var = tk.StringVar(value=str(CONFIG.get("max_record_limit", DEFAULT_CONFIG["max_record_limit"])))
        self.keep_excel_open_var = tk.BooleanVar(value=bool(CONFIG.get("keep_excel_open", True)))
        self.status_var = tk.StringVar(value="就绪")
        self.preview_df = None
        self.preview_config = None

        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.build_ui()

    def build_ui(self):
        self.container = ctk.CTkFrame(self.root, fg_color="#f5f5f7", corner_radius=0)
        self.container.pack(fill="both", expand=True)
        self.container.grid_columnconfigure(1, weight=1)
        self.container.grid_rowconfigure(0, weight=1)

        sidebar = ctk.CTkFrame(self.container, width=72, fg_color="#ffffff", corner_radius=0, border_width=1, border_color="#e5e7eb")
        sidebar.grid(row=0, column=0, sticky="ns")
        sidebar.grid_propagate(False)
        ctk.CTkLabel(sidebar, text="LT", font=ctk.CTkFont(family=UI_FONT, size=18, weight="bold"), text_color="#0a4fd8").pack(pady=(18, 24))
        for icon, active in [("⌂", False), ("▣", True), ("⚙", False), ("ⓘ", False)]:
            ctk.CTkLabel(
                sidebar,
                text=icon,
                width=42,
                height=42,
                fg_color="#0a4fd8" if active else "transparent",
                text_color="#ffffff" if active else "#4b5563",
                corner_radius=10,
                font=ctk.CTkFont(family=UI_FONT, size=18, weight="bold"),
            ).pack(pady=8)

        content = ctk.CTkFrame(self.container, fg_color="#f7f8fb", corner_radius=0)
        content.grid(row=0, column=1, sticky="nsew")
        content.grid_columnconfigure(0, weight=1)
        content.grid_rowconfigure(3, weight=1)
        self.content = content

        hero = ctk.CTkFrame(content, fg_color="#eaf2ff", corner_radius=12, border_width=1, border_color="#d5e3f7")
        hero.grid(row=0, column=0, sticky="ew", padx=24, pady=(22, 16))
        hero.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(
            hero,
            text="↗",
            width=48,
            height=48,
            fg_color="#0a4fd8",
            text_color="#ffffff",
            corner_radius=12,
            font=ctk.CTkFont(family=UI_FONT, size=24, weight="bold"),
        ).grid(row=0, column=0, rowspan=2, padx=(22, 14), pady=22)
        ctk.CTkLabel(
            hero,
            text="巡线记录导入",
            font=ctk.CTkFont(family=UI_FONT, size=25, weight="bold"),
            text_color="#111827",
        ).grid(row=0, column=1, sticky="sw", pady=(22, 0))
        ctk.CTkLabel(
            hero,
            text="自动匹配采集文件，预览后批量追加到主管巡线记录",
            font=ctk.CTkFont(family=UI_FONT, size=13),
            text_color="#4b5563",
        ).grid(row=1, column=1, sticky="nw", pady=(2, 22))
        self.status_badge = ctk.CTkLabel(
            hero,
            textvariable=self.status_var,
            fg_color="#ffffff",
            text_color="#374151",
            corner_radius=8,
            padx=12,
            pady=5,
            font=ctk.CTkFont(family=UI_FONT, size=12, weight="bold"),
        )
        self.status_badge.grid(row=0, column=2, sticky="ne", padx=22, pady=20)

        settings = self.card(content)
        settings.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 14))
        settings.grid_columnconfigure(0, weight=1)
        self.section_title(settings, "本周导入").grid(row=0, column=0, sticky="w", padx=18, pady=(16, 2))
        ctk.CTkLabel(
            settings,
            text="选择来源，计算预览，确认无误后追加到目标表格。",
            font=ctk.CTkFont(family=UI_FONT, size=13),
            text_color="#4b5563",
            anchor="w",
        ).grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))

        form_grid = ctk.CTkFrame(settings, fg_color="transparent")
        form_grid.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 16))
        form_grid.grid_columnconfigure(2, weight=1)
        ctk.CTkLabel(
            form_grid,
            text="来源",
            width=72,
            anchor="w",
            text_color="#111827",
            font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold"),
        ).grid(row=0, column=0, sticky="w", padx=(0, 10), pady=6)
        ctk.CTkSegmentedButton(
            form_grid,
            values=["文件夹", "文件"],
            command=self.change_source_mode,
            variable=self.source_mode_var,
            width=132,
            height=34,
            corner_radius=8,
            selected_color="#0a4fd8",
            selected_hover_color="#0845bd",
            unselected_color="#f8fafc",
            unselected_hover_color="#edf2f7",
            text_color="#111827",
            font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold"),
        ).grid(row=0, column=1, sticky="w", padx=(0, 10), pady=6)

        self.source_path_var = tk.StringVar()
        self.source_entry = ctk.CTkEntry(form_grid, textvariable=self.source_path_var, height=36, corner_radius=8, border_color="#d1d5db", font=ctk.CTkFont(family=UI_FONT, size=13))
        self.source_entry.grid(row=0, column=2, sticky="ew", pady=6)
        self.source_button = ctk.CTkButton(form_grid, text="选择", command=self.choose_active_source, width=86, height=36, corner_radius=8, fg_color="#f8fafc", hover_color="#edf2f7", border_width=1, border_color="#d1d5db", text_color="#111827", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold"))
        self.source_button.grid(row=0, column=3, sticky="e", padx=(8, 0), pady=6)

        ctk.CTkLabel(
            form_grid,
            text="目标文件",
            width=72,
            anchor="w",
            text_color="#111827",
            font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold"),
        ).grid(row=1, column=0, sticky="w", padx=(0, 10), pady=6)
        ctk.CTkLabel(form_grid, text="", width=132).grid(row=1, column=1, sticky="w", padx=(0, 10), pady=6)
        ctk.CTkEntry(form_grid, textvariable=self.target_var, height=36, corner_radius=8, border_color="#d1d5db", font=ctk.CTkFont(family=UI_FONT, size=13)).grid(row=1, column=2, sticky="ew", pady=6)
        ctk.CTkButton(form_grid, text="选择", command=self.choose_target, width=86, height=36, corner_radius=8, fg_color="#f8fafc", hover_color="#edf2f7", border_width=1, border_color="#d1d5db", text_color="#111827", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold")).grid(row=1, column=3, sticky="e", padx=(8, 0), pady=6)
        self.change_source_mode(self.source_mode_var.get())

        actions = ctk.CTkFrame(content, fg_color="transparent")
        actions.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 14))
        actions.grid_columnconfigure(6, weight=1)
        self.preview_button = ctk.CTkButton(actions, text="计算预览", command=self.start_preview, height=38, corner_radius=8, fg_color="#0a4fd8", hover_color="#0845bd", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold"))
        self.preview_button.grid(row=0, column=0, sticky="w")
        self.write_button = ctk.CTkButton(actions, text="确认写入", command=self.confirm_write, height=38, corner_radius=8, fg_color="#16a34a", hover_color="#15803d", state="disabled", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold"))
        self.write_button.grid(row=0, column=1, sticky="w", padx=(10, 0))
        ctk.CTkButton(actions, text="清除", command=self.clear_workspace, height=38, corner_radius=8, fg_color="#f8fafc", hover_color="#edf2f7", border_width=1, border_color="#d1d5db", text_color="#111827", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold")).grid(row=0, column=2, sticky="w", padx=(10, 0))
        ctk.CTkButton(actions, text="配置", command=self.open_config_window, height=38, corner_radius=8, fg_color="#f8fafc", hover_color="#edf2f7", border_width=1, border_color="#d1d5db", text_color="#111827", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold")).grid(row=0, column=3, sticky="w", padx=(10, 0))

        self.preview_card = self.card(content)
        self.preview_card.grid_columnconfigure(0, weight=1)
        self.preview_card.grid_rowconfigure(1, weight=1)
        self.preview_card_visible = False
        self.preview_header = ctk.CTkFrame(self.preview_card, fg_color="transparent")
        self.preview_header.grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 8))
        self.preview_header.grid_columnconfigure(0, weight=1)
        self.preview_title = self.section_title(self.preview_header, "写入前预览")
        self.preview_title.grid(row=0, column=0, sticky="w")
        self.preview_count_label = ctk.CTkLabel(self.preview_header, text="", text_color="#4b5563", font=ctk.CTkFont(family=UI_FONT, size=13))
        self.preview_count_label.grid(row=0, column=1, sticky="e")
        self.preview_list = ctk.CTkScrollableFrame(self.preview_card, fg_color="transparent", height=180)
        self.preview_list.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 16))
        self.preview_card.grid_remove()

    def card(self, parent):
        return ctk.CTkFrame(parent, fg_color="#ffffff", corner_radius=12, border_width=1, border_color="#e5e7eb")

    def section_title(self, parent, text):
        return ctk.CTkLabel(
            parent,
            text=text,
            font=ctk.CTkFont(family=UI_FONT, size=15, weight="bold"),
            text_color="#111827",
        )

    def add_path_row(self, parent, row, label_text, variable, command, button_text="选择"):
        ctk.CTkLabel(parent, text=label_text, text_color="#111827", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold")).grid(row=row, column=0, sticky="w", padx=(0, 10), pady=6)
        ctk.CTkEntry(parent, textvariable=variable, height=36, corner_radius=8, border_color="#d1d5db", font=ctk.CTkFont(family=UI_FONT, size=13)).grid(row=row, column=1, sticky="ew", pady=6)
        ctk.CTkButton(parent, text=button_text, command=command, width=86, height=36, corner_radius=8, fg_color="#f8fafc", hover_color="#edf2f7", border_width=1, border_color="#d1d5db", text_color="#111827", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold")).grid(row=row, column=2, sticky="e", padx=(8, 0), pady=6)

    def add_entry(self, parent, row, column, label_text, variable):
        box = ctk.CTkFrame(parent, fg_color="transparent")
        box.grid(row=row, column=column, sticky="ew", padx=4, pady=5)
        box.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(box, text=label_text, text_color="#111827", anchor="w", font=ctk.CTkFont(family=UI_FONT, size=13, weight="bold")).grid(row=0, column=0, sticky="ew")
        ctk.CTkEntry(box, textvariable=variable, height=36, corner_radius=8, border_color="#d1d5db", font=ctk.CTkFont(family=UI_FONT, size=13)).grid(row=1, column=0, sticky="ew")

    def change_source_mode(self, value):
        mode = "folder" if value in ("folder", "文件夹") else "file"
        self.source_mode_var.set("文件夹" if mode == "folder" else "文件")
        if mode == "folder":
            self.source_path_var.set(self.source_dir_var.get())
            self.source_button.configure(text="选择目录")
        else:
            self.source_path_var.set(self.source_var.get())
            self.source_button.configure(text="选择文件")

    def sync_active_source(self):
        if self.source_mode_var.get() == "文件夹":
            self.source_dir_var.set(self.source_path_var.get().strip())
            self.source_var.set("")
        else:
            self.source_var.set(self.source_path_var.get().strip())

    def choose_active_source(self):
        if self.source_mode_var.get() == "文件夹":
            self.choose_source_dir()
        else:
            self.choose_source()

    def choose_source(self):
        initial = self.source_var.get() or str(BASE_DIR)
        path = filedialog.askopenfilename(
            title="选择采集文件",
            initialdir=str(Path(initial).parent if initial else BASE_DIR),
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        if path:
            self.source_var.set(path)
            self.source_path_var.set(path)

    def choose_target(self):
        initial = self.target_var.get() or str(BASE_DIR)
        path = filedialog.askopenfilename(
            title="选择目标文件",
            initialdir=str(Path(initial).parent if initial else BASE_DIR),
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All Files", "*.*")],
        )
        if path:
            self.target_var.set(path)

    def choose_source_dir(self):
        initial = self.source_dir_var.get() or self.source_var.get() or str(BASE_DIR)
        path = filedialog.askdirectory(
            title="选择采集目录",
            initialdir=str(Path(initial).parent if Path(initial).suffix else Path(initial)),
        )
        if path:
            self.source_dir_var.set(path)
            self.source_path_var.set(path)

    def build_form_config(self):
        self.sync_active_source()
        return {
            "source_file": self.source_var.get().strip(),
            "source_dir": self.source_dir_var.get().strip(),
            "source_keyword": self.source_keyword_var.get().strip() or DEFAULT_CONFIG["source_keyword"],
            "target_file": self.target_var.get().strip(),
            "sheet_name": self.sheet_var.get().strip() or DEFAULT_CONFIG["sheet_name"],
            "target_sheet_name": self.target_sheet_var.get().strip() or DEFAULT_CONFIG["target_sheet_name"],
            "table_name": self.table_name_var.get().strip() or DEFAULT_CONFIG["table_name"],
            "preferred_registrant": self.preferred_var.get().strip() or DEFAULT_CONFIG["preferred_registrant"],
            "preferred_registrant_limit": int(self.preferred_limit_var.get().strip() or DEFAULT_CONFIG["preferred_registrant_limit"]),
            "min_record_limit": int(self.min_limit_var.get().strip() or DEFAULT_CONFIG["min_record_limit"]),
            "max_record_limit": int(self.max_limit_var.get().strip() or DEFAULT_CONFIG["max_record_limit"]),
            "keep_excel_open": bool(self.keep_excel_open_var.get()),
        }

    def open_config_window(self):
        window = ctk.CTkToplevel(self.root)
        window.title("导入配置")
        window.geometry("620x520")
        window.transient(self.root)
        window.grab_set()
        window.configure(fg_color="#f5f5f7")

        panel = self.card(window)
        panel.pack(fill="both", expand=True, padx=18, pady=18)
        panel.grid_columnconfigure((0, 1), weight=1)
        self.section_title(panel, "高级配置").grid(row=0, column=0, columnspan=2, sticky="w", padx=18, pady=(18, 8))

        self.add_entry(panel, 1, 0, "文件关键字", self.source_keyword_var)
        self.add_entry(panel, 1, 1, "源工作表", self.sheet_var)
        self.add_entry(panel, 2, 0, "目标工作表", self.target_sheet_var)
        self.add_entry(panel, 2, 1, "目标表格名", self.table_name_var)
        self.add_entry(panel, 3, 0, "优先登记人", self.preferred_var)
        self.add_entry(panel, 3, 1, "优先数量", self.preferred_limit_var)
        self.add_entry(panel, 4, 0, "最少记录数", self.min_limit_var)
        self.add_entry(panel, 4, 1, "最多记录数", self.max_limit_var)
        ctk.CTkSwitch(
            panel,
            text="完成后保持 Excel 打开",
            variable=self.keep_excel_open_var,
            progress_color="#0a84ff",
            text_color="#374151",
        ).grid(row=5, column=0, columnspan=2, sticky="w", padx=22, pady=(12, 0))

        actions = ctk.CTkFrame(panel, fg_color="transparent")
        actions.grid(row=6, column=0, columnspan=2, sticky="ew", padx=18, pady=(24, 18))
        actions.grid_columnconfigure(0, weight=1)
        ctk.CTkButton(
            actions,
            text="保存",
            command=lambda: self.save_config_window(window),
            height=38,
            corner_radius=19,
            fg_color="#0a84ff",
            hover_color="#006edb",
        ).grid(row=0, column=1, sticky="e")
        ctk.CTkButton(
            actions,
            text="取消",
            command=window.destroy,
            height=38,
            corner_radius=19,
            fg_color="#e5e7eb",
            hover_color="#d1d5db",
            text_color="#111827",
        ).grid(row=0, column=0, sticky="e", padx=(0, 10))

    def save_config_window(self, window):
        try:
            self.persist_form_config()
        except ValueError:
            messagebox.showerror("参数错误", "优先数量、最少记录数和最多记录数必须是整数。")
            return
        window.destroy()

    def persist_form_config(self):
        config = load_config()
        config.update(self.build_form_config())
        save_config(config)
        CONFIG.update(config)
        self.status_var.set("配置已保存")
        logging.info("配置已保存到: %s", CONFIG_PATH)

    def start_preview(self):
        try:
            config = self.build_form_config()
        except ValueError:
            messagebox.showerror("参数错误", "优先数量、最少记录数和最多记录数必须是整数。")
            return

        self.persist_form_config()
        self.clear_preview()
        self.set_busy(True)
        self.write_button.configure(state="disabled")
        self.status_var.set("正在计算预览...")
        logging.info("开始计算写入预览")

        worker = threading.Thread(target=self.preview_in_thread, args=(config,), daemon=True)
        worker.start()

    def preview_in_thread(self, config):
        try:
            result = prepare_transfer(config)
            if result["success"]:
                self.root.after(0, lambda: self.apply_preview_result(result))
            else:
                self.root.after(0, lambda: messagebox.showwarning("未完成", result["message"]))
                self.root.after(0, lambda: self.status_var.set("未找到可导入数据"))
        except Exception as exc:
            logging.exception("预览失败: %s", exc)
            self.root.after(0, lambda: messagebox.showerror("错误", str(exc)))
            self.root.after(0, lambda: self.status_var.set("预览失败"))
        finally:
            self.root.after(0, lambda: self.set_busy(False))

    def apply_preview_result(self, result):
        self.preview_df = result["output_df"]
        self.preview_config = result["config"]
        self.show_preview_card()
        self.refresh_preview_list(self.preview_df)
        self.write_button.configure(state="normal")
        self.status_var.set(f"预览完成，共 {result['record_count']} 条")
        logging.info("预览完成，采集文件: %s", result["source_file"])

    def show_preview_card(self):
        if not self.preview_card_visible:
            self.preview_card.grid(row=3, column=0, sticky="nsew", padx=24, pady=(0, 24))
            self.content.grid_rowconfigure(3, weight=1)
            self.preview_card_visible = True

    def hide_preview_card(self):
        if self.preview_card_visible:
            self.preview_card.grid_remove()
            self.content.grid_rowconfigure(3, weight=0)
            self.preview_card_visible = False

    def refresh_preview_list(self, output_df):
        for child in self.preview_list.winfo_children():
            child.destroy()

        self.preview_count_label.configure(text=f"{len(output_df)} 条待写入")
        for index, row in output_df.reset_index(drop=True).iterrows():
            item = ctk.CTkFrame(self.preview_list, fg_color="#f9fafb", corner_radius=8)
            item.pack(fill="x", pady=3)
            item.grid_columnconfigure(1, weight=1)

            ctk.CTkLabel(
                item,
                text=str(index + 1),
                width=34,
                text_color="#0a4fd8",
                font=ctk.CTkFont(family=UI_FONT, size=12, weight="bold"),
            ).grid(row=0, column=0, sticky="w", padx=(10, 6), pady=7)

            summary = (
                f"{format_preview_date(row.get('日期', ''))}  |  "
                f"{row.get('登记人', '')}  |  "
                f"设备 {row.get('设备编号', '')}  |  "
                f"工单/批次 {row.get('工单号/批次号', '')}  |  "
                f"产品 {row.get('产品图号', '')}  |  "
                f"检查 {row.get('检查正常?', '')}  |  "
                f"问题 {row.get('问题类型', '')}"
            )
            ctk.CTkLabel(
                item,
                text=summary,
                anchor="w",
                font=ctk.CTkFont(family=UI_FONT, size=12),
                text_color="#111827",
            ).grid(row=0, column=1, sticky="ew", padx=(0, 12), pady=7)

    def confirm_write(self):
        if self.preview_df is None or self.preview_df.empty:
            messagebox.showwarning("提示", "请先计算预览。")
            return

        record_count = len(self.preview_df)
        config = self.preview_config or self.build_form_config()
        output_df = self.preview_df.copy()
        self.set_busy(True)
        self.write_button.configure(state="disabled")
        self.status_var.set("正在尝试写入目标文件...")
        logging.info("开始正式写入，共 %s 条", record_count)

        worker = threading.Thread(target=self.write_in_thread, args=(config, output_df), daemon=True)
        worker.start()

    def write_in_thread(self, config, output_df):
        try:
            result = run_transfer(config, prepared_output_df=output_df)
            if result["success"]:
                self.root.after(0, lambda: messagebox.showinfo("完成", result["message"]))
                if result.get("write_mode") == "target":
                    self.root.after(0, lambda: self.status_var.set(f"目标写入完成，共 {result['record_count']} 条"))
                else:
                    self.root.after(0, lambda: self.status_var.set(f"已生成备份，共 {result['record_count']} 条"))
                self.root.after(0, self.clear_preview)
            else:
                self.root.after(0, lambda: messagebox.showwarning("未完成", result["message"]))
                self.root.after(0, lambda: self.status_var.set("写入未完成"))
        except Exception as exc:
            logging.exception("写入失败: %s", exc)
            self.root.after(0, lambda: messagebox.showerror("错误", str(exc)))
            self.root.after(0, lambda: self.status_var.set("写入失败"))
        finally:
            self.root.after(0, lambda: self.set_busy(False))

    def clear_workspace(self):
        self.clear_preview()
        self.status_var.set("已清除，可重新计算")

    def clear_preview(self):
        self.preview_df = None
        self.preview_config = None
        for child in self.preview_list.winfo_children():
            child.destroy()
        self.preview_count_label.configure(text="")
        self.hide_preview_card()
        self.write_button.configure(state="disabled")

    def set_busy(self, is_busy):
        state = "disabled" if is_busy else "normal"
        self.preview_button.configure(state=state)
        if not is_busy and self.preview_df is not None and not self.preview_df.empty:
            self.write_button.configure(state="normal")
        elif is_busy:
            self.write_button.configure(state="disabled")

def run_gui():
    root = ctk.CTk()
    app = TransferApp(root)
    root.mainloop()


if __name__ == "__main__":
    run_gui()
